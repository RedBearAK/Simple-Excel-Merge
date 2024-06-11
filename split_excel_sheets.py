#!/usr/bin/env python3

import os
import sys
import datetime
import openpyxl

# script to split all the tabs (worksheets) of a single Excel spreadsheet file (workbook)
# into separate files in a subfolder

# Fedora packages: 
# - python3-openpyxl

VERBOSE = False
FLUSH = True

def debug(*args, ctx="DD"):
    if not VERBOSE:
        return

    # allow blank lines without context
    if len(args) == 0 or (len(args) == 1 and args[0] == ""):
        print("", flush=FLUSH)
        return
    print(f"({ctx})", *args, flush=FLUSH)

def warn(*args, ctx="WW"):
    print(f"({ctx})", *args, flush=FLUSH)

def error(*args, ctx="EE"):
    print(f"({ctx})", *args, flush=FLUSH)

def log(*args, ctx="--"):
    print(f"({ctx})", *args, flush=FLUSH)

def info(*args, ctx="--"):
    log(*args, ctx=ctx)


def safe_shutdown(exit_code):
    print()
    sys.exit(exit_code)


if len(sys.argv) < 2:
    error("Error: No arguments given. Provide a path to an Excel file with multiple sheets.")
    safe_shutdown(1)

this_file_path                  = os.path.realpath(__file__)
this_file_dir                   = os.path.dirname(this_file_path)
this_file_name                  = os.path.basename(this_file_path)

debug()
debug(f"{this_file_path         = }")
debug(f"{this_file_dir          = }")
debug(f"{this_file_name         = }")

# try:
#     os.makedirs(target_dir_path, exist_ok=True)
# except OSError as os_err:
#     print(f"Error creating directory '{target_dir_path}':\n\t{os_err}")
#     safe_shutdown(1)

print()


def split_worksheets_to_files(excel_file):
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Create a timestamped folder
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    folder_name = f"{os.path.splitext(excel_file)[0]}_{timestamp}"
    os.makedirs(folder_name, exist_ok=True)

    # Iterate through the sheets and save each as a new workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        for row in sheet.iter_rows(values_only=True):
            new_ws.append(row)

        # Save the new workbook
        new_file = os.path.join(folder_name, f"{sheet_name}.xlsx")
        new_wb.save(new_file)
        print(f"Saved: {new_file}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: {this_file_name} <excel_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    if not os.path.isfile(excel_file):
        print(f"Error: File '{excel_file}' not found.")
        sys.exit(1)

    split_worksheets_to_files(excel_file)
